import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup

cidade = 'Morretes'
base_url = 'https://www.accuweather.com/pt/br/morretes/40111/daily-weather-forecast/40111'

headers = {
    'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36"}

list_pagina_dia = 1
list_dia = []
list_temp = []
list_periodo = []
list_nome = []
list_valor_prob_esq = []
list_valor_precip_esq = []
list_valor_precip_dir = []

list_valor_prob= []
list_valor_precip = []
list_valor_horas_prep_dir, list_valor_horas_prep_esq = [], []

while True:
    url_pagina = f'{base_url}?day={list_pagina_dia}'
    site = requests.get(url_pagina, headers=headers)
    soup = BeautifulSoup(site.content, 'html.parser')
    procura_tudo = soup.find_all('div', class_='page-content content-module')

    for item in procura_tudo:
        dia_set = set()
        dia = item.find('span', class_='short-date')
        trata_dia = dia.text.strip()
        dia_set.add(trata_dia)
        print("*" * 30)
        print(f'DATA: {", ".join(dia_set)}')



        procura_temperatura = item.find_all('div', class_='weather')

        for idx, x in enumerate(procura_temperatura): #idx é o indice
            periodo = "DIA" if idx == 0 else "NOITE"
            temperatura = x.find('div', class_='temperature')
            if temperatura:
                trata_temp = temperatura.text.strip()
                trata_temp = trata_temp.replace('Mx', '').replace('Mn', '').strip()
                print(f'{periodo}: Temperatura: {trata_temp}')
                list_periodo.append(periodo)
                list_temp.append(trata_temp)
                list_dia.append(", ".join(dia_set))
            procura_precip_dir = item.find_all('div', class_='right')
            procura_precip_esq = item.find_all('div', class_='left')

            if idx < len(procura_precip_dir):
                itens_precip_dir = procura_precip_dir[idx].find_all('p', class_='panel-item')
                for precip_dir in itens_precip_dir:
                    nome = precip_dir.contents[0].strip()
                    valor = precip_dir.find('span', class_='value').text.strip()
                    print(f'{nome}: {valor}')

                    if nome in 'Precipitação':
                        list_valor_precip.append(valor)



                    if nome in 'Horas de precipitação:':
                        try:
                            valor_tratado = float(valor)
                        except ValueError:
                            valor_tratado = 0
                        list_valor_horas_prep_dir.append(valor_tratado)
                    elif nome in 'Probabilidade de trovoadas':
                        list_valor_horas_prep_dir.append(0)

                    #print(len(itens_precip_dir))

            if idx < len(procura_precip_esq):
                itens_precip_esq = procura_precip_esq[idx].find_all('p', class_='panel-item')
                for precip_esq in itens_precip_esq:
                    nome = precip_esq.contents[0].strip()
                    valor = precip_esq.find('span', class_='value').text.strip()
                    print(f'{nome}: {valor}')
                    if nome in 'Probabilidade de precipitação':
                        list_valor_prob.append(valor)
                    if nome in 'Precipitação':
                        list_valor_precip.append(valor)


            print('_' * 30)

    list_pagina_dia += 1
    if list_pagina_dia == 91:
        break

max_length = max(len(list_dia), len(list_periodo), len(list_temp), len(list_valor_precip),
                 len(list_valor_prob), len(list_valor_horas_prep_dir))

def pad_list(lst, length):
    return lst + [None] * (length - len(lst))

list_dia = pad_list(list_dia, max_length)
list_periodo = pad_list(list_periodo, max_length)
list_temp = pad_list(list_temp, max_length)
list_valor_precip = pad_list(list_valor_precip, max_length)
list_valor_prob = pad_list(list_valor_prob, max_length)
list_valor_horas_prep_dir = pad_list(list_valor_horas_prep_dir, max_length)

dados = {
    'Dia': list_dia,
    'Periodo': list_periodo,
    'Temperatura': list_temp,
    'Precipitação': list_valor_precip,
    'Probabilidade': list_valor_prob,
    'Horas': list_valor_horas_prep_dir
}
df = pd.DataFrame(dados)

file_name = 'dados.xlsx'
df.to_excel(file_name, index=False)

wb = load_workbook(file_name)
ws = wb.active

for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

wb.save(file_name)