import os
import requests
import cairosvg
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

from workbook import formatar_workbook

base_dir = os.path.dirname(os.path.abspath(__file__))
PASTA_SVG = os.path.join(base_dir, 'icones_svg')
PASTA_PNG = os.path.join(base_dir, 'icones_png')
os.makedirs(PASTA_SVG, exist_ok=True)
os.makedirs(PASTA_PNG, exist_ok=True)


headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0 Safari/537.36'
}


def baixar_icone(id_icone):
    svg_url = f"https://www.accuweather.com/images/weathericons/{id_icone}.svg"
    svg_path = os.path.join(PASTA_SVG, f"{id_icone}.svg")
    png_path = os.path.join(PASTA_PNG, f"{id_icone}.png")
    print(f'{svg_path}')
    print(f'{png_path}')
    if not os.path.exists(svg_path):
        try:
            response = requests.get(svg_url, headers=headers, timeout=5)
            if response.status_code == 200:
                with open(svg_path, 'wb') as f:
                    f.write(response.content)
                print(f"Ícone {id_icone} baixado em SVG.")
            else:
                print(f"Erro ao baixar o ícone {id_icone}")
                return None
        except requests.exceptions.RequestException as e:
            print(f"Erro na requisição do ícone {id_icone}: {e}")
            return None


    if not os.path.exists(png_path):
        try:
            cairosvg.svg2png(url=svg_path, write_to=png_path)

            with PILImage.open(png_path) as img:
                img_resized = img.resize((25, 25))
                img_resized.save(png_path)
            print(f"Ícone {id_icone} convertido e redimensionado para PNG.")
        except Exception as e:
            print(f"Erro durante a conversão do {id_icone}: {e}")
            return None

    return png_path if os.path.exists(png_path) else None




def adicionar_icones_ao_excel(df, caminho_arquivo_excel):
    wb = load_workbook(caminho_arquivo_excel)
    ws = wb.active
    ws['E1'] = 'Ícone'
    ws['D1'] = 'Condição'

    # Ajuste de dados nas colunas
    for row in range(2,len(df) +2):
        condicao = ws.cell(row=row, column=5).value
        ws.cell(row=row, column=6).value = condicao
        ws.cell(row=row, column=5).value = None

    for idx, icone_id in enumerate(df['ID'], start=2):
        if icone_id:
            png_path = os.path.join(PASTA_PNG, f"{icone_id}.png")
            if not os.path.exists(png_path):
                png_path = baixar_icone(icone_id)
            if os.path.exists(png_path) and png_path:
                img = Image(png_path)
                cell = ws.cell(row=idx, column=5)
                ws.add_image(img, cell.coordinate)
                print(f"Ícone {icone_id} adicionado ao Excel na linha {idx + 2}.")
            else:
                print(f"PNG para ícone {icone_id} não encontrado.")


    wb.save(caminho_arquivo_excel)
    formatar_workbook(caminho_arquivo_excel)
    print(f"Ícones adicionados ao Excel em: {caminho_arquivo_excel}")